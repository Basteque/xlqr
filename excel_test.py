# -*- coding: utf-8 -*-
"""
Created on Wed Sep  8 14:14:13 2021

@author: basil
"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import qrcode_test as code

workbook = load_workbook(filename="Classeur.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
# print(sheet)
# print(sheet.cell(1,2).value)
# print(code.Logo_link)
sheet.cell(1,3).value = "QR Code"
for i in range(2,39):
    name = str(sheet.cell(i,1).value)
    url = str(sheet.cell(i,2).value)
    img= code.qr_generate(url,name,i)
    # sheet.cell(3,i).value=
    sheet.add_image(img, sheet.cell(i,3))
    
# workbook.save("Classeur_final.xlsx")